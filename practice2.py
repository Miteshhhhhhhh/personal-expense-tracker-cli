#from openpyxl import Workbook, load_workbook

#wb = Workbook()
#ws = wb.active
#ws.title = "Student grades"

#Name = input("Enter your name :")
#Marks = int(input("Enter your marks :"))
#Grades = input("Enter your grades: ")

#ws.append([Name, Marks, Grades])
#wb.save("studentfile.xlsx")

#import os
#os.startfile("studentfile.xlsx")

from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active
ws.append(["Names", "Marks", "Grade"])


while True:
    Name = input("Enter your name :")

    if Name=="done":
        break
    Marks = int(input("Enter your marks:"))
    Grades = input("Enter your grades:")
    ws.append([Name, Marks, Grades])


wb.save("Result.xlsx")

import os
os.startfile("Result.xlsx")



